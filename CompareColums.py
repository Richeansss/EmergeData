import openpyxl

def find_matching_pairs(file1_path, sheet1_name, sheet1_column, file2_path, sheet2_name, key_column_sheet2, columns_to_add):
    # Открываем первый файл Excel
    workbook1 = openpyxl.load_workbook(file1_path)
    sheet1 = workbook1[sheet1_name]

    # Открываем второй файл Excel
    workbook2 = openpyxl.load_workbook(file2_path)
    sheet2 = workbook2[sheet2_name]

    # Получаем данные из столбцов и создаем словарь для значений второй таблицы
    sheet1_data = [(cell.row, cell.value.strip().lower() if isinstance(cell.value, str) else cell.value) for cell in sheet1[sheet1_column]]
    sheet2_data = {sheet2[key_column_sheet2 + str(row)].value.strip().lower() if isinstance(sheet2[key_column_sheet2 + str(row)].value, str) else sheet2[key_column_sheet2 + str(row)].value: [sheet2[column + str(row)].value for column in columns_to_add] for row in range(1, sheet2.max_row + 1)}

    # Находим индекс последнего заголовка в первой таблице
    last_header_column_index = None
    for col_index, cell in enumerate(sheet1[1], start=1):
        if cell.value is not None:
            last_header_column_index = col_index

    if last_header_column_index is None:
        raise ValueError("No headers found in sheet1 starting from row 3")

    # Определяем, с какого столбца начинать добавление в первую таблицу
    start_column_index = last_header_column_index + 1

    # Добавляем указанные столбцы из второй таблицы в первую и собираем пары
    matching_pairs = []
    for row1, value1 in sheet1_data:
        if value1 in sheet2_data:
            values_from_sheet2 = sheet2_data[value1]
            matching_pairs.append((row1, value1, values_from_sheet2))
            for col_index, column_name in enumerate(columns_to_add, start=start_column_index):
                sheet1.cell(row=row1, column=col_index).value = values_from_sheet2[col_index - start_column_index]

    # Находим значения из второй таблицы, которые не были добавлены в первую таблицу
    values_not_matched = []
    for value2 in sheet2_data:
        if all(value2 != row[1] for row in matching_pairs):  # Проверяем, что значение из sheet2_data не было добавлено в sheet1
            values_not_matched.append((value2, sheet2_data[value2]))

    # Выводим значения, которые не попали в пары из второй таблицы
    for value2, values_from_sheet2 in values_not_matched:
        print(f"Value in Sheet2 not matched in Sheet1: {value2}, Columns: {values_from_sheet2}")

    # Сохраняем изменения
    workbook1.save(file1_path)

    print(f"Total matching pairs found: {len(matching_pairs)}")

    return matching_pairs

# Пример использования
file1_path = 'C:\\Users\\Alexey\\IdeaProjects\\DataMerge\\ДашбордСОВОФ.xlsx'
sheet1_name = 'База График ТОиР и МТР'
sheet1_column = 'B'
file2_path = 'C:\\Users\\Alexey\\IdeaProjects\\DataMerge\\SortedData_Output.xlsx'
sheet2_name = 'GroupedData_Summary'
key_column_sheet2 = 'AC'
columns_to_add = ['BO', 'BP', 'BQ', 'BR', 'BS', 'AO', 'T']  # Список столбцов для добавления из второй таблицы

matching_pairs = find_matching_pairs(file1_path, sheet1_name, sheet1_column, file2_path, sheet2_name, key_column_sheet2, columns_to_add)

for row1, value1, values_from_sheet2 in matching_pairs:
    print(f"Matching pair found: Sheet1, Row {row1}: {value1} == Sheet2: {values_from_sheet2}")
