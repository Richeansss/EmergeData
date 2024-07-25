import openpyxl
from datetime import datetime
import sys
from ExtractRows import create_summary_rows

# Устанавливаем кодировку для стандартного ввода и вывода
sys.stdout.reconfigure(encoding='utf-8')
sys.stdin.reconfigure(encoding='utf-8')

def convert_dates_in_place(file_path, sheet_name, columns):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]
    except Exception as e:
        print(f"Ошибка при загрузке файла или листа: {e}")
        return

    header_row = sheet[1]
    column_indices = {}
    for col in columns:
        for cell in header_row:
            if cell.value and col in str(cell.value):
                column_indices[col] = cell.col_idx
                break

    for col_name, col_idx in column_indices.items():
        for row in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    try:
                        if isinstance(cell.value, str):
                            try:
                                cell.value = datetime.strptime(cell.value, "%d.%m.%Y %H:%M:%S").strftime("%d.%м.%Y")
                            except ValueError:
                                try:
                                    cell.value = datetime.strptime(cell.value, "%d.%m.%Y").strftime("%d.%m.%Y")
                                except ValueError:
                                    continue
                        elif isinstance(cell.value, datetime):
                            cell.value = cell.value.strftime("%d.%m.%Y")

                        cell.number_format = 'dd.mm.yyyy'
                    except ValueError:
                        continue

    wb.save(file_path)
    print(f"Даты успешно преобразованы и сохранены в файл '{file_path}'")

def remove_leading_apostrophe(file_path, sheet_name, column_name):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]
    except Exception as e:
        print(f"Ошибка при загрузке файла или листа: {e}")
        return

    header_row = sheet[1]
    column_index = None
    for cell in header_row:
        if cell.value and column_name in str(cell.value):
            column_index = cell.col_idx
            break

    if column_index is None:
        print(f"Не найден столбец с именем '{column_name}'")
        return

    for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("'"):
                cell.value = cell.value[1:]

    wb.save(file_path)
    print(f"Апострофы успешно удалены в столбце '{column_name}' и сохранены в файл '{file_path}'")

# Пример использования функции
file_path = 'C:\\Users\\Alexey\\IdeaProjects\\DataMerge\\MergedData.xlsx'
sheet_name = 'GroupedData'
columns = ['Дата отгрузки по графику (from file1)', 'Дата начала работ (from file2)', 'Дата окончания работ (from file2)']

convert_dates_in_place(file_path, sheet_name, columns)

# Удаляем апострофы
remove_leading_apostrophe(file_path, sheet_name, 'ППП (from file2)')

# Вызов функции из extract_row.py
start_marker = 'Код позиции (from file1)'
end_marker = 'Кол-во'
shipped_column = 'Статус поставки на склад СУ Сургут (да/нет) (from file1)'
schedule_date_column = 'Поставлено / ориентировочная дата поставки на объект (from file1)'  # Надо исправить
start_work_column = 'Дата начала работ (from file2)'
end_work_column = 'Дата окончания работ (from file2)'

summary_rows = create_summary_rows(file_path, sheet_name, start_marker, end_marker, shipped_column,
                                   schedule_date_column, start_work_column, end_work_column)
print("Обработка завершена.")
