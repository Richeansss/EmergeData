import openpyxl

def move_ppp_column_to_left(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb['GroupedData_Summary']

        # Найти индекс столбца с заголовком "ППП"
        ppp_column_index = None
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value and "ППП" in str(cell.value):
                ppp_column_index = col_idx
                break

        if ppp_column_index is None:
            print("Столбец с заголовком 'ППП' не найден или пустой.")
            return None

        # Проверка, что столбец не пустой
        all_non = True
        for row_idx in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=ppp_column_index).value
            if cell_value is not None and cell_value != "Non":
                all_non = False
                break

        if all_non:
            print("Столбец с заголовком 'ППП' содержит только 'Non'. Нет необходимости перемещать.")
            return None

        # Вычислить количество столбцов, которые нужно передвинуть
        col_shift = ppp_column_index - 1

        if col_shift <= 0:
            print("Столбец с заголовком 'ППП' уже в крайнем левом положении.")
            return None

        # Вставить новый столбец перед столбцом с заголовком "ППП"
        sheet.insert_cols(1, amount=1)

        # Скопировать данные из старого столбца в новый
        for row_idx in range(2, sheet.max_row + 1):
            old_cell = sheet.cell(row=row_idx, column=ppp_column_index)
            new_cell = sheet.cell(row=row_idx, column=1)
            new_cell.value = old_cell.value

        # Удалить старый столбец
        sheet.delete_cols(ppp_column_index)

        # Сохранить изменения
        wb.save(file_path)
        print("Столбец с заголовком 'ППП' успешно перемещен в крайнее левое положение.")

    except Exception as e:
        print(f"Ошибка при обработке файла: {e}")

