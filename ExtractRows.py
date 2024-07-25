import openpyxl
import sys
from datetime import datetime

# Устанавливаем кодировку для стандартного ввода и вывода
sys.stdout.reconfigure(encoding='utf-8')
sys.stdin.reconfigure(encoding='utf-8')

def load_workbook_and_sheet(file_path, sheet_name):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]
        return wb, sheet
    except Exception as e:
        print(f"Ошибка при загрузке файла или листа: {e}")
        return None, None

def find_column_index(header_row, column_name):
    for cell in header_row:
        if cell.value and column_name in str(cell.value):
            return cell.col_idx - 1
    return None

def parse_date(date_str):
    try:
        return datetime.strptime(date_str, "%d.%m.%Y")
    except ValueError:
        return None

def find_latest_date(sheet, start_row, end_row, col_index):
    latest_date = None
    for row_idx in range(start_row, end_row + 1):
        cell_value = sheet.cell(row=row_idx, column=col_index + 1).value
        if isinstance(cell_value, str):
            cell_value = parse_date(cell_value)
        if latest_date is None or (cell_value and cell_value > latest_date):
            latest_date = cell_value
    return latest_date

def find_ranges(sheet, start_marker, end_marker):
    ranges = []
    start_row = None
    first_row_skipped = False
    for row in sheet.iter_rows():
        if not first_row_skipped:
            first_row_skipped = True
            continue

        for cell in row:
            if cell.value and start_marker in str(cell.value):
                start_row = cell.row + 1
                start_value = cell.value
            if cell.value and end_marker in str(cell.value):
                end_row = cell.row - 1
                if start_row is not None:
                    ranges.append((start_row, end_row, start_value))
                start_row = None
                break
    return ranges

def count_yes_no_in_shipped_column(sheet, ranges, shipped_column_index):
    results = []
    total_yes_count = 0
    total_no_count = 0

    for start_row, end_row, start_value in ranges:
        yes_count = 0
        no_count = 0
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
            cell_value = row[shipped_column_index].value
            if isinstance(cell_value, str):
                cell_value_lower = cell_value.lower()
                if cell_value_lower == "да":
                    yes_count += 1
                elif cell_value_lower == "нет":
                    no_count += 1
        results.append((start_value, start_row, end_row, yes_count, no_count))
        total_yes_count += yes_count
        total_no_count += no_count

    return results, total_yes_count, total_no_count

def create_summary_rows(file_path, sheet_name, start_marker, end_marker, shipped_column, schedule_date_column,
                        start_work_column, end_work_column):
    wb, sheet = load_workbook_and_sheet(file_path, sheet_name)
    if not wb or not sheet:
        return None

    new_sheet_name = f"{sheet_name}_Summary"
    if new_sheet_name in wb.sheetnames:
        wb.remove(wb[new_sheet_name])
    new_sheet = wb.create_sheet(new_sheet_name)

    ranges = find_ranges(sheet, start_marker, end_marker)
    if not ranges:
        print("Не удалось найти диапазоны между маркерами.")
        return None

    header_row = sheet[1]
    shipped_col_index = find_column_index(header_row, shipped_column)
    schedule_date_col_index = find_column_index(header_row, schedule_date_column)
    start_work_col_index = find_column_index(header_row, start_work_column)
    end_work_col_index = find_column_index(header_row, end_work_column)

    errors = []
    if shipped_col_index is None:
        errors.append(f"Не удалось найти столбец '{shipped_column}'")
    if schedule_date_col_index is None:
        errors.append(f"Не удалось найти столбец '{schedule_date_column}'")
    if start_work_col_index is None:
        errors.append(f"Не удалось найти столбец '{start_work_column}'")
    if end_work_col_index is None:
        errors.append(f"Не удалось найти столбец '{end_work_column}'")

    if errors:
        print("Ошибка: " + ", ".join(errors))
        return None

    header_row_values = [cell.value for cell in header_row]
    summary_header = header_row_values + ['Последняя дата отгрузки', 'Последняя дата начала работ',
                                          'Последняя дата окончания работ', 'Да', 'Нет']
    summary_rows = [summary_header]

    results, total_yes_count, total_no_count = count_yes_no_in_shipped_column(sheet, ranges, shipped_col_index)

    for start_row, end_row, start_value in ranges:
        summary_row = [None] * sheet.max_column
        for col_idx in range(sheet.max_column):
            values_in_range = set()
            for row_idx in range(start_row, end_row + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx + 1).value
                values_in_range.add(cell_value)
            summary_row[col_idx] = next(iter(values_in_range)) if len(values_in_range) == 1 else "Non"

        latest_ship_date = find_latest_date(sheet, start_row, end_row, schedule_date_col_index)
        latest_start_work_date = find_latest_date(sheet, start_row, end_row, start_work_col_index)
        latest_end_work_date = find_latest_date(sheet, start_row, end_row, end_work_col_index)
        summary_row.extend([latest_ship_date, latest_start_work_date, latest_end_work_date])

        for start_val, row_start, row_end, yes_count, no_count in results:
            if start_value == start_val:
                summary_row.extend([yes_count, no_count])

        summary_rows.append(summary_row)

    for col_idx, header in enumerate(summary_rows[0]):
        new_sheet.cell(row=1, column=col_idx + 1, value=header)

    for row_idx, row_data in enumerate(summary_rows[1:], start=2):
        for col_idx, value in enumerate(row_data):
            new_sheet.cell(row=row_idx, column=col_idx + 1, value=value)

    wb.save(file_path)
    return summary_rows
